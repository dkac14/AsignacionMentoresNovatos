from openpyxl import load_workbook, Workbook
from collections import defaultdict, Counter
import os


# Mentores
COL_M_NOMBRE = "Nombre"
COL_M_CARRERA = "CARRERA2"
COL_M_CORREO = "Correo electrónico"
COL_M_TELEF = "Número Telefónico"

# Novatos
COL_N_NOMBRE = "Nombre Completo"
COL_N_CARRERA = "CARRERA2"
COL_N_CORREO = "Correo de Espol, como estudiante de Espol te debieron asignar un correo, si todavia no sabes cual es, puedes dejarlo en blanco"
COL_N_TELEF = "Número Telefónico"

# ======= RUTAS DE LOS ARCHIVOS DE ENTRADA =======
ruta_mentores = r"C:\Users\Lenovo\Downloads\PAO II 2025 Inscripcion Mentores.xlsx"
ruta_novatos  = r"C:\Users\Lenovo\Downloads\PAO II 2025 Inscripcion Novatos.xlsx"

# ======= UTILIDADES =======
def norm(s):
    
    if s is None:
        return ""
    return str(s).strip().upper()

def indices_por_nombre(encabezados, cols):
    
    idxs = []
    for c in cols:
        try:
            idxs.append(encabezados.index(c))
        except ValueError:
            raise ValueError(f"No se encontró la columna '{c}'. Encabezados: {encabezados}")
    return idxs

def leer_tabla(ws, cols):
    
    headers = [cell.value for cell in ws[1]]
    idxs = indices_por_nombre(headers, cols)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {}
        for key, idx in zip(cols, idxs):
            item[key] = row[idx]
        rows.append(item)
    return rows


wbM = load_workbook(ruta_mentores); wsM = wbM.active
wbN = load_workbook(ruta_novatos);  wsN = wbN.active

mentores = leer_tabla(wsM, [COL_M_NOMBRE, COL_M_CARRERA, COL_M_CORREO, COL_M_TELEF])
novatos  = leer_tabla(wsN, [COL_N_NOMBRE, COL_N_CARRERA, COL_N_CORREO, COL_N_TELEF])


mentores_por_carrera = defaultdict(list)
for m in mentores:
    key = norm(m[COL_M_CARRERA])
    mentores_por_carrera[key].append(m)

novatos_por_carrera = defaultdict(list)
for n in novatos:
    key = norm(n[COL_N_CARRERA])
    novatos_por_carrera[key].append(n)


asignaciones = []   
sin_mentor   = []  
sin_novatos  = []   

for carrera, lista_nov in novatos_por_carrera.items():
    lista_ment = mentores_por_carrera.get(carrera, [])
    if not lista_ment:
        
        for n in lista_nov:
            sin_mentor.append((carrera, n))
        continue

    
    m_count = len(lista_ment)
    for i, n in enumerate(lista_nov):
        mentor = lista_ment[i % m_count]
        asignaciones.append((carrera, mentor, n))


for carrera, lista_ment in mentores_por_carrera.items():
    if carrera not in novatos_por_carrera or len(novatos_por_carrera[carrera]) == 0:
        for m in lista_ment:
            sin_novatos.append((carrera, m))


conteo_por_mentor = Counter()
for _, mentor, _ in asignaciones:
    clave = (mentor[COL_M_NOMBRE], mentor[COL_M_CORREO], norm(mentor[COL_M_CARRERA]))
    conteo_por_mentor[clave] += 1


wb_out = Workbook()
ws_asig = wb_out.active
ws_asig.title = "Asignaciones"
ws_res  = wb_out.create_sheet("Resumen")
ws_sinM = wb_out.create_sheet("Novatos_sin_mentor")
ws_sinN = wb_out.create_sheet("Mentores_sin_novatos")


headers_asig = [
    "CARRERA",
    "Mentor - Nombre", "Mentor - Correo", "Mentor - Teléfono",
    "Novato - Nombre", "Novato - Correo", "Novato - Teléfono"
]
ws_asig.append(headers_asig)

for carrera, mentor, novato in asignaciones:
    ws_asig.append([
        carrera,
        mentor.get(COL_M_NOMBRE), mentor.get(COL_M_CORREO), mentor.get(COL_M_TELEF),
        novato.get(COL_N_NOMBRE), novato.get(COL_N_CORREO), novato.get(COL_N_TELEF)
    ])


ws_res.append(["Mentor - Nombre", "Mentor - Correo", "Carrera", "Total Novatos"])
for (m_nombre, m_correo, carrera), cnt in sorted(conteo_por_mentor.items(), key=lambda x: (x[0][2], x[0][0])):
    ws_res.append([m_nombre, m_correo, carrera, cnt])


ws_res.append([])
ws_res.append(["Carrera", "Mentores", "Novatos", "Asignaciones"])
for carrera in sorted(set(list(mentores_por_carrera.keys()) + list(novatos_por_carrera.keys()))):
    ment_ct = len(mentores_por_carrera.get(carrera, []))
    nov_ct  = len(novatos_por_carrera.get(carrera, []))
    asig_ct = sum(1 for c, _, _ in asignaciones if c == carrera)
    ws_res.append([carrera, ment_ct, nov_ct, asig_ct])


ws_sinM.append(["Carrera", "Novato - Nombre", "Novato - Correo", "Novato - Teléfono"])
for carrera, n in sin_mentor:
    ws_sinM.append([carrera, n.get(COL_N_NOMBRE), n.get(COL_N_CORREO), n.get(COL_N_TELEF)])


ws_sinN.append(["Carrera", "Mentor - Nombre", "Mentor - Correo", "Mentor - Teléfono"])
for carrera, m in sin_novatos:
    ws_sinN.append([carrera, m.get(COL_M_NOMBRE), m.get(COL_M_CORREO), m.get(COL_M_TELEF)])


salida = "Asignaciones.xlsx"
wb_out.save(salida)
print(f"Archivo generado: {os.path.abspath(salida)}")

